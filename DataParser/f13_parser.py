import os
import pandas as pd
import openpyxl


def extract_data(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    data_frames = {
        'vš celkem': [],
        'veřejné vš celkem': [],
        'soukromé vš celkem': []
    }

    # Check all lists
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        year = sheet_name
        current_category = None
        bold_flag = 0  # Check for bold rows

        for row in sheet.iter_rows(min_row=9, min_col=4, max_col=11, values_only=False):
            if not any(cell.value for cell in row):
                break  # End of the table
            if row[0].font.bold:
                bold_flag += 1
                if bold_flag == 1:
                    current_category = 'vš celkem'
                elif bold_flag == 2:
                    current_category = 'veřejné vš celkem'
                elif bold_flag == 3:
                    current_category = 'soukromé vš celkem'
                continue

            if current_category:
                data_frames[current_category].append([cell.value for cell in row] + [year])

    # Table setup
    columns = ['Město', 'Celkem', 'Bakalářské', 'Magisterské', 'Doktorské', 'Poprvé zapsaní', 'Absolventi', 'Cizinci',
               'Rok']

    category_data_frames = {
        category: pd.DataFrame(data, columns=columns)
        for category, data in data_frames.items()
    }

    return category_data_frames


def process_data(data_frames, base_path='mesta_rok_'):
    if not os.path.exists(base_path):
        os.makedirs(base_path)

    for category, df in data_frames.items():
        for year in df['Rok'].unique():
            year_df = df[df['Rok'] == year]
            filename = f"{base_path}/{category.replace(' ', '_')}_data_{year}.xlsx"
            year_df.to_excel(filename, index=False)
            print(f"Data for {year} saved to {filename}")


# Read data
data_frames = extract_data('f13.xlsx')

# Process the data
process_data(data_frames)
