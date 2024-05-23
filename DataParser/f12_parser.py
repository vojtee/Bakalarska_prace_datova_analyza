import os
import pandas as pd
import openpyxl

def extract_data(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    all_data = []
    output_folder = 'univerzity_rok_'

    # Check if the output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    private_data = []
    public_data = []

    for sheet_name in workbook.sheetnames:
        if int(sheet_name) >= 2020 and int(sheet_name) <= 2022:
            sheet = workbook[sheet_name]
            rows = process_data(sheet, sheet_name)
            df = pd.DataFrame(rows, columns=columns)

            # Extract public and private school data
            private_row, public_row = extract_public_private(sheet, sheet_name)
            private_data.append(private_row)
            public_data.append(public_row)

            # Saving data to separate Excel sheet
            df.to_excel(f"{output_folder}/{sheet_name}.xlsx", index=False)

            # Combined data Excel sheet
            all_data.append(df)

    # Saving all data to one Excel sheet
    final_df = pd.concat(all_data, ignore_index=True)
    final_df.to_excel("univerzity.xlsx", index=False)
    print("Data saved successfully to univerzity.xlsx")

    # Saving public and private school data to separate Excel sheets
    private_df = pd.DataFrame(private_data, columns=columns)
    public_df = pd.DataFrame(public_data, columns=columns)

    private_df.to_excel("soukrome_skoly.xlsx", index=False)
    public_df.to_excel("verejne_skoly.xlsx", index=False)
    print("Data saved successfully to soukrome_skoly.xlsx and verejne_skoly.xlsx")


def process_data(sheet, year):
    rows = []
    is_before_2019 = int(year) <= 2018  # Check for years before 2019 to skip "1. lékařská fakulta" which is bold in the table
    for row in sheet.iter_rows(min_row=10, min_col=4, max_col=17, max_row=sheet.max_row):
        if row[0].font.bold:
            # the check for the "1. lékařská fakulta"
            if is_before_2019 and '1. lékařská fakulta' in (row[0].value or ''):
                continue
            row_data = [row[0].value] + [cell.value for cell in row[3:]] + [year]
            rows.append(row_data)
    return rows


def extract_public_private(sheet, year):
    # Extract data from row 9 (combined) and row 10 (public) for columns G to Q
    combined_row = [int(sheet.cell(row=9, column=col).value) for col in range(7, 18)]
    public_row = [int(sheet.cell(row=10, column=col).value) for col in range(7, 18)]

    # Calculate private school data as the difference between combined and public
    private_row = [combined_row[i] - public_row[i] for i in range(len(combined_row))]

    # Append year to the rows
    private_row.append(year)
    public_row.append(year)

    # Add school type as a placeholder
    private_row.insert(0, "Soukromé školy")
    public_row.insert(0, "Veřejné školy")

    return private_row, public_row


# Table setup
columns = ['Škola', 'Celkem', 'Celkem ženy', 'Studující celkem ČR', 'Poprvé zapsaní ČR', 'Absolventi ČR',
           'Studující celkem ciz', 'Poprvé zapsaní ciz', 'Absolventi ciz', 'Přerušené studium', 'Jinoplatci',
           'Samoplatci', 'Rok']

# Process the data
extract_data('f12.xlsx')
